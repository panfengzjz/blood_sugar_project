#coding: utf-8
import openpyxl
import time

def ward_classify(ward):
    res = ""
    if (ward=="四十七病区" or ward=="八病区"or ward=="二病区"or ward=="九病区"or ward=="六病区"or ward=="七病区"or ward=="三病区"or 
        ward=="三十七病区"or ward=="三十三病区"or ward=="三十四病区"or ward=="三十五病区"or ward=="十八病区"or ward=="十病区"or ward=="十二病区"or 
        ward=="十六病区"or ward=="十七病区"or ward=="十三病区"or ward=="十四病区"or ward=="十一病区"or ward=="四病区"or ward=="四十八病区"or ward=="四十九病区"or ward=="四十六病区"
        or ward=="四十七病区"or ward=="四十四病区" or ward=="外科日间病部" or ward=="五病区产科" or ward=="五病区妇科" or ward=="五十一病区"):
        res="wai"
    elif (ward=="二十八病区" or ward=="二十病区"or ward=="二十九病区"or ward=="二十七病区"or ward=="二十一病区"or ward=="六十八病区"or ward=="六十病区"or ward=="六十九病区"
          or ward=="廿六病区"or ward=="廿四病区"or ward=="廿五病区"or ward=="三十八病区"or ward=="三十病区" or ward=="三十二病区A" or ward=="三十二病区B" or ward=="三十九病区" or ward=="三十六病区" or ward=="三十一病区" or ward=="十九病区" or ward=="四十病区" or ward=="四十二病区" or ward=="四十三病区" or ward=="四十一病区" or ward=="五十病区" or ward=="五十二病区" or ward=="五十六病区" or ward=="五十三病区" or ward=="一病区" or ward=="周转三部" or ward=="周转四部"):
        res="nei"
    elif (ward=="（东院）肝外监护室" or ward=="（东院）心外监护室"or ward=="（东院）心脏内科监护室"or ward=="呼吸内科监护室" or ward=="急诊ICU" or ward=="神经内科监护室" or ward=="外科监护室" or ward=="外科监护室A "    ):
        res="jian"
    elif (ward=="二十三病区" ):
        res="endo"
    return res

# src_name: 原始表格名称
# ret_name: 最终生成的新excel名称
def ward_exchange(src_name, ret_name1,ret_name2,ret_name3,ret_name4):
    wb2 = openpyxl.Workbook()
    wb2.save(ret_name1)
    wb3 = openpyxl.Workbook()
    wb3.save(ret_name2)
    wb4 = openpyxl.Workbook()
    wb4.save(ret_name3)   
    wb5 = openpyxl.Workbook()
    wb5.save(ret_name4)    
    print('新建成功')

    wb1 = openpyxl.load_workbook(src_name)
    wb2 = openpyxl.load_workbook(ret_name1)
    wb3 = openpyxl.load_workbook(ret_name2)
    wb4 = openpyxl.load_workbook(ret_name3)
    wb5 = openpyxl.load_workbook(ret_name4)
    
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]
    sheet3 = wb3[wb3.sheetnames[0]]
    sheet4 = wb4[wb4.sheetnames[0]]
    sheet5 = wb5[wb5.sheetnames[0]]
    max_row = sheet1.max_row        #最大行数
    max_column = sheet1.max_column  #最大列数
    
    i = 2
    n1=2
    n2=2
    n3=2
    n4=2
    
    while(i < max_row+1):
        ward=sheet1.cell(i, 5).value
        ward_class=ward_classify(ward)
        
        if (ward_class=="wai"):
            for col in range(1, 15):
                sheet2.cell(n1, col).value = sheet1.cell(i, col).value             
            n1+=1  
        elif (ward_class=="nei"):
            for col in range(1, 15):
                sheet3.cell(n2, col).value = sheet1.cell(i, col).value
                
            n2+=1   
        elif (ward_class=="jian"):
            for col in range(1, 15):
                sheet4.cell(n3, col).value = sheet1.cell(i, col).value             
            n3+=1  
        elif (ward_class=="endo"):
            #print("endo")
            for col in range(1, 15):
                sheet5.cell(n4, col).value = sheet1.cell(i, col).value             
            n4+=1
            #print(n4)
        i += 1


    wb2.save(ret_name1)
    wb3.save(ret_name2)
    wb4.save(ret_name3)
    wb5.save(ret_name4)
    #保存数据
    wb1.close()
    wb2.close()
    wb3.close()
    wb4.close()
    wb5.close()
    print("功能完成")

if __name__ == "__main__":
    start=time.time()
    src_name = "new-huge.xlsx"
    ret_name1 = "wai.xlsx"   
    ret_name2 = "nei.xlsx"  
    ret_name3 = "jian.xlsx" 
    ret_name4="endo.xlsx"
    #patientList = make_list(fileName)   #这次生成一个数组，方便pop
    #reorder_excel(fileName, reslName)
    ward_exchange(src_name, ret_name1,ret_name2,ret_name3,ret_name4)
    print(time.time()-start)