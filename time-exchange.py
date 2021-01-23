#coding: utf-8
import openpyxl
#import datetime

# global parameter
patientList = []   # 保存源excel中的姓名

# 将住院号写入数组并返回
def make_list(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]
    new_list = []
    max_row = sheet.max_row

    for i in range(1, max_row+1):
        item = sheet.cell(i, 3).value  # 名字在第三列
        if item not in new_list:
            new_list.append(item)
    return new_list

# src_name: 原始表格名称
# ret_name: 最终生成的新excel名称
def reorder_excel(src_name, ret_name):
    wb2 = openpyxl.Workbook()
    wb2.save(ret_name)
    print('新建成功')

    wb1 = openpyxl.load_workbook(src_name)
    wb2 = openpyxl.load_workbook(ret_name)
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]
    max_row = sheet1.max_row        #最大行数
    max_column = sheet1.max_column  #最大列数
    row_count = 1                   #保存当前 ret_name excel 中记录到的行数

    while(patientList):
        target_name = patientList.pop(0)        #当前循环需要找的名字
        for m in range(1, max_row+1):
            cur_flow_name = sheet1.cell(m, 3).value
            if (cur_flow_name != target_name):  #如果名字不同，则找下一行
                continue
            for n in range(1, max_column+1):    #如果相同则写入 ret_name 的下一行中
                cell1 = sheet1.cell(m, n).value #获取data单元格数据
                sheet2.cell(row_count, n).value = cell1 #赋值到test单元格
            row_count += 1

    wb2.save(ret_name)   #保存数据
    wb1.close()
    wb2.close()

# 计算excel中读出的时间与所需时间点中的哪个最接近
def calc_time_diff(excel_time):
    if (type(excel_time) == str):
        return ""
    hour = excel_time.hour
    mins = excel_time.minute
    time_num = hour*60 + mins
    res = ""
    if (4*60 <= time_num < 7*60+30):        # from 4:00~7:30
        res = "6点"
    elif (7*60+30 <= time_num < 9*60+30):   # from 7:30~9:30
        res = "8点30"
    elif (9*60+30 <= time_num < 12*60):     # from 9:30~12:00
        res = "10点30"
    elif (12*60 <= time_num < 15*60+30):    # from 12:00~15:30
        res = "13点"
    elif (15*60+30 <= time_num < 18*60):    # from 15:30~18:00
        res = "16点30"
    elif (18*60 <= time_num < 20*60):       # from 18:00~20:00
        res = "19点"
    elif (20*60 <= time_num < 24*60):       # from 20:00~24:00
        res = "21点"
    return res

# src_name: 原始表格名称
# ret_name: 最终生成的新excel名称
def time_exchange(src_name, ret_name):
    wb2 = openpyxl.Workbook()
    wb2.save(ret_name)
    print('新建成功')

    wb1 = openpyxl.load_workbook(src_name)
    wb2 = openpyxl.load_workbook(ret_name)
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]
    max_row = sheet1.max_row        #最大行数
    max_column = sheet1.max_column  #最大列数
    
    i = 1
    while(i < max_row+1):
        excel_time = sheet1.cell(i, 9).value
        time_diff = calc_time_diff(excel_time)
        sheet2.cell(i, 9).value = excel_time
        sheet2.cell(i, 14).value = time_diff
        for col in range(1, 9):
            sheet2.cell(i, col).value = sheet1.cell(i, col).value         
        i += 1


    wb2.save(ret_name)   #保存数据
    wb1.close()
    wb2.close()
    print("功能完成")

if __name__ == "__main__":
    fileName = "test.xlsx"
    reslName = "new-huge.xlsx"   #两次跑结果会覆盖
    #patientList = make_list(fileName)   #这次生成一个数组，方便pop
    #reorder_excel(fileName, reslName)
    time_exchange(fileName, reslName)
